import pandas as pd
from dateutil import parser
from datetime import datetime
import arrow
import string
import warnings
from typing import List, Dict, Tuple
from django.http import JsonResponse
import numpy as np
import pandas as pd
import os
from openpyxl import load_workbook
from rest_framework.response import Response

# Utility function to convert date column with automatic method selection

def infer_dtypes(values: List, sample_size: int = 300, stop_after: int = 300) -> str:
    """
    Infers the data type by randomly sampling from a list. Values are explicitly converted to string before checking.
    Args:
        values (list): A list to infer data types from.
        sample_size (int, optional): The number of values to sample from the list. Entire list will be sampled if set to None. Defaults to 300.
        stop_after (int, optional): The maximum number of non-empty values needed for the test. Equal to sample_size if set to None. Defaults to 300.
    Returns:
        str: The inferred data type ('number', 'bool', 'str', 'mixed', 'empty', 'categorical', 'date').
    """
    found = 0
    non_empty_count = 0

    sample_size = sample_size if sample_size is not None else len(values)
    stop_after = stop_after if stop_after is not None else sample_size

    for v in np.random.choice(values, sample_size):
        v = str(v)
        if v != '':
            non_empty_count += 1
            if non_empty_count > stop_after:
                break
            try:
                # Check for integers or floats
                float(v)  # This will catch both integers and floats
                found |= 3  # Use 3 to represent 'number' (covers both int and float)
            except ValueError:
                if v.lower() in ['true', 'false']:
                    found |= 4
                else:
                    try:
                        # Check for dates
                        pd.to_datetime(v, errors='raise')
                        found |= 8
                    except (ValueError, TypeError):
                        found |= 16

    # Check if the data is mixed
    if bin(found).count('1') > 1:
        return 'mixed'

    if found & 16:
        return 'str'
    elif found & 8:
        return 'date'
    elif found & 4:
        return 'bool'
    elif found & 3:  # This represents 'number'
        return 'number'
    else:
        return 'empty'


def infer_and_convert_data_types(df: pd.DataFrame, column, data_type) -> pd.DataFrame:
    """
    A helper function to infer and convert data types.
    
    Tries to convert columns to numeric, datetime, or categorical types.
    
    Args:
        df (pd.DataFrame): The DataFrame to infer and convert data types for.
        
    Returns:
        pd.DataFrame: The DataFrame with converted data types.
    """
    for col in df.columns:
        # Attempt to infer the type of the column
        inferred_type = infer_dtypes(df[col])
        if(column == col):
            inferred_type = data_type
        print(f'Column: {col}, Inferred Type: {inferred_type}')

        # Convert based on inferred type
        if inferred_type == 'number':
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        elif inferred_type == 'complex':
            df[col] = df[col].apply(lambda x: complex(x) if pd.notna(x) else np.nan)
        elif inferred_type == 'bool':
            df[col] = df[col].astype('bool')
        elif inferred_type == 'str':
            df[col] = df[col].astype('string').fillna('')
        elif inferred_type in ['datetime', 'date']:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        elif inferred_type == 'mixed':
            # If mixed, convert to a categorical type if the number of unique values is less than half
            if df[col].nunique() / len(df[col]) < 0.5:
                df[col] = pd.Categorical(df[col])

    return df


import os  # Ensure to import os at the beginning of your file

def process_file(file_path,column, data_type):
    try:
        processed_data = []  # Initialize a list to store processed data
        data_types = {}  # Initialize a dictionary to store data types
        
        # Extract the filename from the file path
        filename = os.path.basename(file_path)
        print(filename)
        # Read the file based on its type
        if filename.endswith('.csv'):
            # Read CSV file in chunks
            for chunk in pd.read_csv(file_path, chunksize=1000, low_memory=False):
                # Infer and convert data types for the chunk
                chunk = infer_and_convert_data_types(chunk, column, data_type)
                processed_data.append(chunk.to_dict(orient='records'))

                # Capture the data types from the first chunk
                if not data_types:
                    data_types = chunk.dtypes.apply(lambda x: x.name).to_dict()

        elif filename.endswith('.xlsx'):
            # Read Excel file using openpyxl
            workbook = load_workbook(filename=file_path, read_only=True)  # Load the workbook
            sheet = workbook.active  # Get the active sheet

            # Read the data in chunks (e.g., 1000 rows at a time)
            chunk_size = 1000
            data_chunk = []
            for row in sheet.iter_rows(values_only=True):
                data_chunk.append(row)
                if len(data_chunk) == chunk_size:
                    # Convert to DataFrame and process
                    df_chunk = pd.DataFrame(data_chunk[1:], columns=data_chunk[0])  # Use the first row as headers
                    df_chunk = infer_and_convert_data_types(df_chunk,column, data_type)
                    processed_data.append(df_chunk.to_dict(orient='records'))

                    # Capture the data types from the first chunk
                    if not data_types:
                        data_types = df_chunk.dtypes.apply(lambda x: x.name).to_dict()

                    data_chunk = []  # Reset for the next chunk

            # Process any remaining rows in the last chunk
            if data_chunk:
                df_chunk = pd.DataFrame(data_chunk[1:], columns=data_chunk[0])  # Use the first row as headers
                df_chunk = infer_and_convert_data_types(df_chunk,column, data_type)
                processed_data.append(df_chunk.to_dict(orient='records'))

                # Capture the data types from the last chunk if not captured already
                if not data_types:
                    data_types = df_chunk.dtypes.apply(lambda x: x.name).to_dict()

        else:
            return JsonResponse({'error': 'Unsupported file type'}, status=400)

        # Flatten the list of processed data
        flat_processed_data = [item for sublist in processed_data for item in sublist]

        # Return both the data and its types
        return Response({'data': flat_processed_data, 'data_types': data_types, 'file_path': file_path}, status=200)

    except Exception as e:
        print(f"Error processing file: {e}")  # Log the error for debugging
        return JsonResponse({'error': 'Error processing file'}, status=500)

